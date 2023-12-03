from tkinter import *
from tkinter.ttk import Combobox
import  tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from openpyxl.utils import get_column_letter

root = Tk()
root.title("B2-Prüfung Analyse")
root.geometry('500x600+300+200')
root.resizable(False, False)
root.configure(bg='#6970F5')

file = pathlib.Path('Formular_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] = "Vorname"
    sheet['B1'] = "Nachname"
    sheet['C1'] = "Alter"
    sheet['D1'] = "Geschlecht"
    sheet['E1'] = "Staatsangehörigkeit"
    sheet['F1'] = "Anzahl der Besuche"
    sheet['G1'] = "Prüfungsstatus"

    file.save('Formular_data.xlsx')



def submit():
    vorname = vornameValue.get()
    nachname = nachnameValue.get()
    age = ageValue.get()
    gender = gender_combobox.get()
    nationality = nationalValue.get()
    num_visits = num_visitsValue.get()
    exam_pass = exam_combobox.get()

    file = openpyxl.load_workbook('Formular_data.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value = vorname)
    sheet.cell(column=2, row=sheet.max_row, value = nachname)
    sheet.cell(column=3, row=sheet.max_row, value = age)
    sheet.cell(column=4, row=sheet.max_row, value = gender)
    sheet.cell(column=5, row=sheet.max_row, value = nationality)
    sheet.cell(column=6, row=sheet.max_row, value = num_visits)
    sheet.cell(column=7, row=sheet.max_row, value = exam_pass)
    file.save(r'Formular_data.xlsx')
    clear()
    messagebox.showinfo('Datentransfer', 'Ihre Daten wurden erfolgreich übertragen.\n Vielen Dank für Ihre Aufmerksamkeit.')

def clear():
    vornameValue.set('')
    nachnameValue.set('')
    ageValue.set('')
    nationalValue.set('')
    num_visitsValue.set('')




#apps icon
icon_image = PhotoImage(file="analysis.png")
root.iconphoto(False,icon_image)

#heading
Label(root, text="Bitte füllen Sie dieses Anmeldeformular aus:", font="arial 13", bg="#6970F5", fg="#fff").place(x=20,y=20)

#label
Label(root, text="Vorname:", font=23, bg="#6970F5", fg="#fff").place(x=50,y=100)
Label(root, text="Nachname:", font=23, bg="#6970F5", fg="#fff").place(x=50,y=150)
Label(root, text="Alter:", font=23, bg="#6970F5", fg="#fff").place(x=50,y=200)
Label(root, text="Geschlecht:", font=23, bg="#6970F5", fg="#fff").place(x=50,y=250)
Label(root, text="Staatsangehörigkeit:", font=23, bg="#6970F5", fg="#fff").place(x=50,y=300)
Label(root, text="Anzahl der Besuche:", font=23, bg="#6970F5", fg="#fff").place(x=50,y=350)
Label(root, text="Prüfungsstatus:", font=23, bg="#6970F5", fg="#fff").place(x=50,y=400)

#Entry
vornameValue = StringVar()
nachnameValue = StringVar()
ageValue = StringVar()
nationalValue = StringVar()
num_visitsValue = StringVar()


vornameEntry = Entry(root, textvariable=vornameValue, width=19, bd=2, font=20)
nachnameEntry = Entry(root, textvariable=nachnameValue, width=19, bd=2, font=20)
ageEntry = Entry(root, textvariable=ageValue, width=19, bd=2, font=20)
nationalEntry = Entry(root, textvariable=nationalValue, width=11, bd=2, font=20)
num_visitsEntry = Entry(root, textvariable=num_visitsValue, width=11, bd=2, font=20)

#Geschlecht
gender_combobox = Combobox(root, values= ["männlich", "weiblich", "divers"], font= 'arial 14', state='r', width=17)
gender_combobox.place(x=165, y=250)
gender_combobox.set("männlich")

#Prüfungsstatus
exam_combobox = Combobox(root, values= ["bestanden", "nicht bestanden"], font= 'arial 14', state='r', width=13)
exam_combobox.place(x=206, y=400)

vornameEntry.place(x=160, y=100)
nachnameEntry.place(x=160, y=150)
ageEntry.place(x=160, y=200)
nationalEntry.place(x=245, y=300)
num_visitsEntry.place(x=245, y=350)


Button(root, text='Bestätigen', bg='#7BA1F4', fg='white', width=15, height=2, command=submit).place(x=190,y=500)
Button(root, text='Löschen', bg='#7BA1F4', fg='white', width=15, height=2, command=clear).place(x=40,y=500)
Button(root, text='Beenden', bg='#7BA1F4', fg='white', width=15, height=2, command= lambda: root.destroy()).place(x=340,y=500)


root.mainloop()